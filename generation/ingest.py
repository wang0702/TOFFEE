
from __future__ import annotations

import csv as _csv
import hashlib
import logging
import re
import sqlite3
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from toffee.core.executor import execute_sql, list_tables

log = logging.getLogger(__name__)


_KEY_NAME_PATTERNS = re.compile(
    r"(?:^id$|_id$|^sku$|_sku$|^code$|_code$|^date$|_date$|^region$|"
    r"^customer$|_customer$|^product$|_product$|^order$|_order$|^key$|_key$)",
    re.IGNORECASE,
)


_TOKEN_STOP = {"the", "of", "and", "or", "to", "a", "an", "in", "for", "by"}

_TOKEN_RE = re.compile(r"[A-Za-z][A-Za-z0-9]*")


def _tokenize(text: str) -> List[str]:
    return [t.lower() for t in _TOKEN_RE.findall(text or "") if t.lower() not in _TOKEN_STOP]


@dataclass
class SourceUnit:
    unit_id: str
    format: str
    title: str
    logical_table: str
    origin_path: str
    origin_locator: str
    db_path: str
    sch: List[Tuple[str, str, bool]] = field(default_factory=list)
    stat: Dict[str, Any] = field(default_factory=dict)
    key: List[str] = field(default_factory=list)
    ent: List[str] = field(default_factory=list)


    scratch_db_path: str = ""

    def col_tokens(self) -> List[str]:
        out: List[str] = []
        for col, _t, _n in self.sch:
            out.extend(_tokenize(col))
        return out

    def has_num(self) -> bool:
        return any(_is_num_type(t) for _, t, _ in self.sch)

    def has_date(self) -> bool:
        return any(_is_date_col(c, t) for c, t, _ in self.sch)


def _is_num_type(t: str) -> bool:
    t = (t or "").upper()
    return any(k in t for k in ("INT", "REAL", "FLOAT", "DOUBLE", "NUM", "DEC"))


def _is_date_col(col: str, t: str) -> bool:
    if "DATE" in (t or "").upper() or "TIME" in (t or "").upper():
        return True
    return bool(re.search(r"date|time|year|month|day", col or "", re.IGNORECASE))


_PERCENT_NAME_RE = re.compile(r"pct|percent|rate|ratio|share", re.IGNORECASE)
_MONEY_NAME_RE = re.compile(
    r"price|cost|revenue|amount|salary|spend|budget|fee|usd|dollar",
    re.IGNORECASE,
)
_COUNT_NAME_RE = re.compile(
    r"count|number|qty|quantity|^num_|_num$", re.IGNORECASE)


def _col_role(col: str, typ: str, st: Dict[str, Any]) -> str:
    dr = float(st.get("distinct_ratio", 0.0) or 0.0)
    if _is_date_col(col, typ):
        return "time"
    if _KEY_NAME_PATTERNS.search(col or "") and (_is_num_type(typ) or dr >= 0.5):
        return "identifier"
    if dr >= 0.95:
        return "identifier"
    if _is_num_type(typ):
        return "measure"
    if 0.0 < dr < 0.5:
        return "category"
    return "entity"


def _value_kind(col: str, typ: str, st: Dict[str, Any], role: str) -> str:
    if role == "identifier":
        return "id"
    if role == "measure":
        if _PERCENT_NAME_RE.search(col or ""):
            return "percent"
        lo, hi = st.get("min"), st.get("max")
        if lo is not None and hi is not None and 0.0 <= float(lo) and float(hi) <= 1.0:
            return "percent"
        if _MONEY_NAME_RE.search(col or ""):
            return "monetary"
        if _COUNT_NAME_RE.search(col or ""):
            return "count"
    return "free"


_MAX_ENT_TOKENS = 32
_STAT_SAMPLE_ROWS = 500
_TOP_CATEGORIES = 3


_SAMPLE_VALUES_N = 20


def _profile_sqlite_table(db_path: str, table: str) -> Tuple[
    List[Tuple[str, str, bool]], Dict[str, Any], List[str], List[str]
]:
    conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)


    import time as _t
    _deadline = _t.monotonic() + float(__import__("os").environ.get("TOFFEE_PROFILE_BUDGET_S", "120"))
    conn.set_progress_handler(lambda: 1 if _t.monotonic() > _deadline else 0, 100_000)
    try:
        cur = conn.cursor()
        cur.execute(f'PRAGMA table_xinfo("{table}");')
        pragma = [
            row for row in cur.fetchall()
            if len(row) < 7 or int(row[6] or 0) != 1
        ]
        sch: List[Tuple[str, str, bool]] = [
            (row[1], row[2] or "", not bool(row[3])) for row in pragma
        ]
        if not sch:
            return sch, {}, [], []

        try:
            cur.execute(f'SELECT COUNT(*) FROM "{table}";')
            n_rows = int(cur.fetchone()[0])
        except Exception:
            n_rows = 0

        per_col_stats: Dict[str, Dict[str, Any]] = {}
        key_by_distinct: List[str] = []
        top_values: List[str] = []

        sample_n = min(_STAT_SAMPLE_ROWS, n_rows) if n_rows > 0 else 0
        for col, typ, _null in sch:
            st: Dict[str, Any] = {"null_rate": 0.0, "distinct_ratio": 0.0}
            if n_rows > 0:
                try:
                    cur.execute(
                        f'SELECT COUNT(*) - COUNT("{col}") FROM "{table}";'
                    )
                    nulls = int(cur.fetchone()[0])
                    st["null_rate"] = nulls / max(1, n_rows)
                except Exception:
                    pass
                try:
                    cur.execute(
                        f'SELECT COUNT(DISTINCT "{col}") FROM '
                        f'(SELECT "{col}" FROM "{table}" LIMIT {sample_n});'
                    )
                    distinct = int(cur.fetchone()[0])
                    st["distinct_ratio"] = distinct / max(1, sample_n)
                except Exception:
                    pass
                if _is_num_type(typ):
                    try:
                        cur.execute(
                            f'SELECT AVG(CAST("{col}" AS REAL) * CAST("{col}" AS REAL)) '
                            f'- AVG(CAST("{col}" AS REAL)) * AVG(CAST("{col}" AS REAL)), '
                            f'MIN(CAST("{col}" AS REAL)), MAX(CAST("{col}" AS REAL)) '
                            f'FROM (SELECT "{col}" FROM "{table}" WHERE "{col}" IS NOT NULL LIMIT {sample_n});'
                        )
                        var, lo, hi = cur.fetchone()
                        st["variance"] = float(var) if var is not None else 0.0
                        if lo is not None:
                            st["min"] = float(lo)
                        if hi is not None:
                            st["max"] = float(hi)
                    except Exception:
                        st["variance"] = 0.0
                    try:
                        cur.execute(
                            f'SELECT DISTINCT "{col}" FROM '
                            f'(SELECT "{col}" FROM "{table}" WHERE "{col}" IS NOT NULL '
                            f'LIMIT {sample_n}) LIMIT {_SAMPLE_VALUES_N};'
                        )
                        st["sample_values"] = [str(r[0]) for r in cur.fetchall()]
                    except Exception:
                        pass
                else:
                    try:
                        cur.execute(
                            f'SELECT "{col}", COUNT(*) c FROM "{table}" '
                            f'WHERE "{col}" IS NOT NULL GROUP BY "{col}" '
                            f'ORDER BY c DESC LIMIT {_TOP_CATEGORIES};'
                        )
                        tops = [str(r[0]) for r in cur.fetchall()]
                        if tops:
                            st["top_categories"] = tops
                            top_values.extend(tops)
                    except Exception:
                        pass

            if st.get("distinct_ratio", 0.0) >= 0.95 and n_rows > 0:
                key_by_distinct.append(col)
            st["role"] = _col_role(col, typ, st)
            st["value_kind"] = _value_kind(col, typ, st, st["role"])
            per_col_stats[col] = st

        key_by_name = [c for c, _, _ in sch if _KEY_NAME_PATTERNS.search(c or "")]
        key = list(dict.fromkeys(key_by_name + key_by_distinct))

        stat = {
            "n_rows": n_rows,
            "n_cols": len(sch),
            "has_date": any(_is_date_col(c, t) for c, t, _ in sch),
            "has_num": any(_is_num_type(t) for _, t, _ in sch),
            "cols": per_col_stats,
        }

        ent_tokens: List[str] = list(dict.fromkeys(
            _tokenize(table)
            + [tok for c, _, _ in sch for tok in _tokenize(c)]
            + [tok for v in top_values for tok in _tokenize(v)]
        ))[:_MAX_ENT_TOKENS]

        return sch, stat, key, ent_tokens
    finally:
        conn.close()


def _uid(*parts: str) -> str:
    h = hashlib.sha256("||".join(parts).encode()).hexdigest()
    return h[:10]


def _safe_table_name(stem: str, kind: str, *parts: str) -> str:
    base = re.sub(r"[^A-Za-z0-9_]", "_", stem or "").strip("_").lower() or kind
    if base[0].isdigit():
        base = f"t_{base}"
    return f"{base}_{_uid(kind, *parts)[:6]}"


def _qident(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'


_CREATE_TABLE_HEAD = re.compile(
    r'^\s*CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?'
    r'(?:(?:"(?:[^"]|"")*")|(?:`[^`]*`)|(?:\[[^\]]*\])|(?:[^\s(]+))',
    re.IGNORECASE,
)


def _copy_sqlite_table(
    db_path: str, source_table: str, scratch_db: str, target_table: str,
) -> bool:
    """Copy one native SQLite table into the environment scratch database.

    Task construction and trajectory search must query the same database.  In
    particular, a path may combine a native SQLite table with a CSV, sheet, or
    Markdown table that already lives in ``scratch_db``.  Copying the native
    table once at ingestion gives every later component one stable namespace.
    """
    conn = sqlite3.connect(scratch_db)
    try:
        conn.execute("ATTACH DATABASE ? AS source_db", (db_path,))
        pragma = conn.execute(
            f"PRAGMA source_db.table_xinfo({_qident(source_table)})"
        ).fetchall()
        if not pragma:
            return False
        materialized = [
            row for row in pragma
            if len(row) < 7 or int(row[6] or 0) != 1
        ]
        inserted = [
            row for row in materialized
            if len(row) < 7 or int(row[6] or 0) == 0
        ]
        ddl_row = conn.execute(
            "SELECT type, sql FROM source_db.sqlite_master WHERE name = ?",
            (source_table,),
        ).fetchone()
        conn.execute(f"DROP TABLE IF EXISTS main.{_qident(target_table)}")
        copied_ddl = False
        if ddl_row and ddl_row[0] == "table" and ddl_row[1]:
            rewritten, n_sub = _CREATE_TABLE_HEAD.subn(
                f"CREATE TABLE main.{_qident(target_table)}",
                str(ddl_row[1]), count=1,
            )
            if n_sub == 1:
                try:
                    conn.execute(rewritten)
                    copied_ddl = True
                except sqlite3.Error as exc:
                    log.debug("DDL copy failed for %s::%s; materializing values: %s",
                              db_path, source_table, exc)
        if not copied_ddl:
            col_defs: List[str] = []
            for row in materialized:
                col = str(row[1])
                typ = str(row[2] or "")
                not_null = " NOT NULL" if bool(row[3]) else ""
                col_defs.append(f"{_qident(col)} {typ}{not_null}".strip())
            conn.execute(
                f"CREATE TABLE main.{_qident(target_table)} ({', '.join(col_defs)})"
            )
            inserted = materialized
        cols = [_qident(str(row[1])) for row in inserted]
        if not cols:
            return False
        col_list = ", ".join(cols)
        conn.execute(
            f"INSERT INTO main.{_qident(target_table)} ({col_list}) "
            f"SELECT {col_list} FROM source_db.{_qident(source_table)}"
        )
        conn.commit()
        return True
    except Exception as exc:
        try:
            conn.execute(f"DROP TABLE IF EXISTS main.{_qident(target_table)}")
            conn.commit()
        except Exception:
            pass
        log.warning("SQLite materialization failed for %s::%s: %s",
                    db_path, source_table, exc)
        return False
    finally:
        try:
            conn.execute("DETACH DATABASE source_db")
        except Exception:
            pass
        conn.close()


def _source_foreign_keys(db_path: str, table: str) -> List[Dict[str, str]]:
    try:
        conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
        try:
            rows = conn.execute(
                f"PRAGMA foreign_key_list({_qident(table)})"
            ).fetchall()
        finally:
            conn.close()
    except Exception:
        return []
    return [
        {"table": str(r[2]), "from": str(r[3]), "to": str(r[4])}
        for r in rows if len(r) >= 5 and r[2] and r[3] and r[4]
    ]


def _ingest_sqlite(
    db_path: str, scratch_db: str, used_tables: Optional[Set[str]] = None,
) -> List[SourceUnit]:
    result = list_tables(db_path)
    if not result.success:
        return []
    tables = [t.strip() for t in result.stdout.split() if t.strip()]
    used_tables = used_tables if used_tables is not None else set()

    # Allocate every destination name before copying so foreign-key targets can
    # be remapped consistently when two database files contain the same table.
    table_map: Dict[str, str] = {}
    for tbl in tables[:40]:
        target = _reserve_table_name(
            tbl, used_tables, "sqlite", db_path, tbl,
        )
        table_map[tbl] = target

    units: List[SourceUnit] = []
    for tbl in tables[:40]:
        target = table_map[tbl]
        if not _copy_sqlite_table(db_path, tbl, scratch_db, target):
            continue
        try:
            sch, stat, key, ent = _profile_sqlite_table(scratch_db, target)
        except Exception as exc:
            log.warning("Profile failed for %s :: %s : %s", db_path, tbl, exc)
            continue
        if not sch:
            continue
        stat["declared_fks"] = [
            {**fk, "table": table_map.get(fk["table"], fk["table"])}
            for fk in _source_foreign_keys(db_path, tbl)
        ]
        units.append(SourceUnit(
            unit_id=_uid("sqlite_table", db_path, tbl),
            format="sqlite_table",
            title=tbl,
            logical_table=target,
            origin_path=db_path,
            origin_locator=f"{db_path}::{tbl}",
            db_path=scratch_db,
            sch=sch, stat=stat, key=key, ent=ent,
            scratch_db_path=scratch_db,
        ))
    return units


def _safe_ident(name: str) -> str:
    ident = re.sub(r"[^A-Za-z0-9_]", "_", name or "")
    return ident or "col"


def _reserve_table_name(
    proposed: str, used_tables: Set[str], kind: str, *parts: str,
) -> str:
    used_lower = {name.lower() for name in used_tables}
    target = proposed
    if target.lower() in used_lower:
        target = _safe_table_name(
            f"{proposed}_{kind}", f"{kind}_collision", *parts,
        )
        base = target
        suffix = 1
        while target.lower() in used_lower:
            target = f"{base}_{suffix}"
            suffix += 1
    used_tables.add(target)
    return target


def _coerce_sqlite_type(values: List[str]) -> str:
    seen_int = seen_float = seen_text = False
    for v in values:
        if v is None or v == "":
            continue
        try:
            int(v)
            seen_int = True
            continue
        except ValueError:
            pass
        try:
            float(v)
            seen_float = True
            continue
        except ValueError:
            pass
        seen_text = True
        break
    if seen_text:
        return "TEXT"
    if seen_float:
        return "REAL"
    if seen_int:
        return "INTEGER"
    return "TEXT"


def _materialize_rows_into_sqlite(
    scratch_db: str, table: str, header: List[str], rows: List[List[Any]]
) -> Tuple[List[Tuple[str, str, bool]], int]:
    header = [_safe_ident(h) for h in header] or ["col"]

    seen: Dict[str, int] = {}
    unique_header: List[str] = []
    for h in header:
        if h in seen:
            seen[h] += 1
            unique_header.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            unique_header.append(h)

    col_values: List[List[str]] = [[] for _ in unique_header]
    normalized: List[List[Any]] = []
    for r in rows:
        if not r:
            continue
        r = list(r) + [None] * (len(unique_header) - len(r))
        r = r[: len(unique_header)]
        normalized.append(r)
        for i, v in enumerate(r):
            if v is None:
                continue
            col_values[i].append(str(v))
    if not normalized:
        return [], 0

    types = [_coerce_sqlite_type(col_values[i]) for i in range(len(unique_header))]
    sch = [(unique_header[i], types[i], True) for i in range(len(unique_header))]

    conn = sqlite3.connect(scratch_db)
    try:
        cur = conn.cursor()
        cols_def = ", ".join(f'"{h}" {t}' for h, t in zip(unique_header, types))
        cur.execute(f'DROP TABLE IF EXISTS "{table}";')
        cur.execute(f'CREATE TABLE "{table}" ({cols_def});')
        placeholders = ",".join("?" * len(unique_header))
        cur.executemany(
            f'INSERT INTO "{table}" VALUES ({placeholders});',
            [tuple(r) for r in normalized],
        )
        conn.commit()
    finally:
        conn.close()
    return sch, len(normalized)


def _ingest_csv(
    file_path: str, scratch_db: str, used_tables: Optional[Set[str]] = None,
) -> List[SourceUnit]:
    p = Path(file_path)
    if not p.is_file():
        return []
    sep = "\t" if p.suffix.lower() == ".tsv" else ","
    try:
        with p.open("r", encoding="utf-8", errors="replace", newline="") as fh:
            reader = _csv.reader(fh, delimiter=sep)
            rows = list(reader)
    except Exception as exc:
        log.warning("CSV read failed for %s: %s", file_path, exc)
        return []
    if len(rows) < 2:
        return []
    header, data = rows[0], rows[1:]
    used_tables = used_tables if used_tables is not None else set()
    table = _reserve_table_name(
        _safe_table_name(p.stem, "csv", file_path),
        used_tables, "csv", file_path,
    )
    sch, n_rows = _materialize_rows_into_sqlite(scratch_db, table, header, data)
    if not sch:
        return []
    _sch, stat, key, ent = _profile_sqlite_table(scratch_db, table)
    return [SourceUnit(
        unit_id=_uid("csv_table", file_path),
        format="csv_table",
        title=p.stem,
        logical_table=table,
        origin_path=file_path,
        origin_locator=file_path,
        db_path=scratch_db,
        sch=_sch or sch, stat=stat, key=key, ent=ent,
        scratch_db_path=scratch_db,
    )]


def _ingest_excel(
    file_path: str, scratch_db: str, used_tables: Optional[Set[str]] = None,
) -> List[SourceUnit]:
    p = Path(file_path)
    if not p.is_file():
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except ImportError:
        log.warning("openpyxl not installed; skipping %s", file_path)
        return []
    except Exception as exc:
        log.warning("Excel read failed for %s: %s", file_path, exc)
        return []

    units: List[SourceUnit] = []
    used_tables = used_tables if used_tables is not None else set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]

        all_rows = [r for r in all_rows if any(c is not None and str(c) != "" for c in r)]
        if len(all_rows) < 2:
            continue
        header = [str(c) if c is not None else "col" for c in all_rows[0]]
        data = [list(r) for r in all_rows[1:]]
        table = _reserve_table_name(
            _safe_table_name(
                f"{p.stem}_{sheet_name}", "xlsx", file_path, sheet_name,
            ),
            used_tables, "xlsx", file_path, sheet_name,
        )
        sch, n_rows = _materialize_rows_into_sqlite(scratch_db, table, header, data)
        if not sch:
            continue
        _sch, stat, key, ent = _profile_sqlite_table(scratch_db, table)
        units.append(SourceUnit(
            unit_id=_uid("excel_sheet", file_path, sheet_name),
            format="excel_sheet",
            title=f"{p.stem}::{sheet_name}",
            logical_table=table,
            origin_path=file_path,
            origin_locator=f"{file_path}::{sheet_name}",
            db_path=scratch_db,
            sch=_sch or sch, stat=stat, key=key, ent=ent,
            scratch_db_path=scratch_db,
        ))
    wb.close()
    return units


_MD_PIPE_ROW = re.compile(r"^\s*\|(.+)\|\s*$")
_MD_SEP_ROW = re.compile(r"^\s*\|?\s*:?-{3,}:?(\s*\|\s*:?-{3,}:?)*\s*\|?\s*$")
_MD_KV_LINE = re.compile(r"^\s*([A-Za-z][A-Za-z0-9_ ]{0,40}?)\s*:\s*(.+?)\s*$")
_MD_HEADER = re.compile(r"^(#{1,6})\s+(.+?)\s*$")


def _extract_md_pipe_tables(text: str) -> List[Tuple[List[str], List[List[str]]]]:
    lines = text.splitlines()
    out: List[Tuple[List[str], List[List[str]]]] = []
    i = 0
    while i < len(lines):
        if _MD_PIPE_ROW.match(lines[i]) and i + 1 < len(lines) and _MD_SEP_ROW.match(lines[i + 1]):
            header = [c.strip() for c in lines[i].strip().strip("|").split("|")]
            rows: List[List[str]] = []
            j = i + 2
            while j < len(lines) and _MD_PIPE_ROW.match(lines[j]):
                cells = [c.strip() for c in lines[j].strip().strip("|").split("|")]
                if len(cells) == len(header):
                    rows.append(cells)
                j += 1
            if len(rows) >= 2:
                out.append((header, rows))
            i = j
        else:
            i += 1
    return out


def _extract_md_kv_blocks(text: str) -> List[Tuple[List[str], List[List[str]]]]:
    lines = text.splitlines()

    triples: List[Tuple[str, str, str]] = []
    current_section = ""
    records: List[Dict[str, str]] = []
    current: Dict[str, str] = {}

    for ln in lines:
        hm = _MD_HEADER.match(ln)
        if hm:
            if current:
                records.append(current)
                current = {}
            current_section = hm.group(2).strip()
            continue
        m = _MD_KV_LINE.match(ln)
        if m:
            k = _safe_ident(m.group(1))
            v = m.group(2)
            triples.append((current_section or "_root", k, v))
            if k in current:
                if current:
                    records.append(current)
                current = {}
            current[k] = v
        elif ln.strip() == "":
            if current:
                records.append(current)
                current = {}
    if current:
        records.append(current)

    out: List[Tuple[List[str], List[List[str]]]] = []

    if len(triples) >= 3:
        rows = [[s, k, v] for s, k, v in triples]
        out.append((["section", "key", "value"], rows))

    if len(records) >= 3:
        key_count: Dict[str, int] = {}
        for rec in records:
            for k in rec:
                key_count[k] = key_count.get(k, 0) + 1
        kept_keys = [k for k, c in key_count.items() if c >= 3]
        if len(kept_keys) >= 2:
            stable_records = [r for r in records if all(k in r for k in kept_keys)]
            if len(stable_records) >= 2:
                rows = [[r.get(k, "") for k in kept_keys] for r in stable_records]
                out.append((kept_keys, rows))

    return out


def _ingest_md(
    file_path: str, scratch_db: str, used_tables: Optional[Set[str]] = None,
) -> List[SourceUnit]:
    p = Path(file_path)
    if not p.is_file():
        return []
    try:
        text = p.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:
        log.warning("MD read failed for %s: %s", file_path, exc)
        return []
    blocks = _extract_md_pipe_tables(text) + _extract_md_kv_blocks(text)
    units: List[SourceUnit] = []
    used_tables = used_tables if used_tables is not None else set()
    for idx, (header, data) in enumerate(blocks):
        table = _reserve_table_name(
            _safe_table_name(f"{p.stem}_{idx}", "md", file_path, str(idx)),
            used_tables, "md", file_path, str(idx),
        )
        sch, n_rows = _materialize_rows_into_sqlite(scratch_db, table, header, data)
        if not sch:
            continue
        _sch, stat, key, ent = _profile_sqlite_table(scratch_db, table)
        units.append(SourceUnit(
            unit_id=_uid("md_struct", file_path, str(idx)),
            format="md_struct",
            title=f"{p.stem}#{idx}",
            logical_table=table,
            origin_path=file_path,
            origin_locator=f"{file_path}#{idx}",
            db_path=scratch_db,
            sch=_sch or sch, stat=stat, key=key, ent=ent,
            scratch_db_path=scratch_db,
        ))
    return units


def _detect_family(path: str) -> Optional[str]:
    low = path.lower()
    if low.endswith(".sqlite") or low.endswith(".db") or low.endswith(".sqlite3"):
        return "sqlite_table"
    if low.endswith(".csv") or low.endswith(".tsv"):
        return "csv_table"
    if low.endswith(".xlsx") or low.endswith(".xls") or low.endswith(".xlsm"):
        return "excel_sheet"
    if low.endswith(".md") or low.endswith(".markdown"):
        return "md_struct"
    return None


_SCRATCH_APP_ID = 0x544F4646  # "TOFF"


def _owned_scratch(path: Path) -> bool:
    try:
        conn = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
        try:
            row = conn.execute("PRAGMA application_id").fetchone()
            return bool(row and int(row[0]) == _SCRATCH_APP_ID)
        finally:
            conn.close()
    except Exception:
        return False


def _managed_scratch_location(path: Path) -> bool:
    return (
        path.name.startswith("env_")
        and path.parent.name in {".toffee_scratch", "toffee_scratch"}
    )


def ingest_environment(env_sources: List[str], scratch_db: str) -> List[SourceUnit]:
    scratch_parent = Path(scratch_db).parent
    scratch_parent.mkdir(parents=True, exist_ok=True)

    scratch_path = Path(scratch_db)
    source_paths = set()
    for src in env_sources:
        try:
            source_paths.add(Path(src).resolve())
        except Exception:
            continue
    if scratch_path.resolve() in source_paths:
        raise ValueError("scratch_db must be different from every source database")
    if scratch_path.exists():
        if not _owned_scratch(scratch_path) and not _managed_scratch_location(scratch_path):
            raise ValueError("refusing to replace a scratch database not owned by TOFFEE")
        scratch_path.unlink()
    scratch_conn = sqlite3.connect(scratch_db)
    try:
        scratch_conn.execute(f"PRAGMA application_id = {_SCRATCH_APP_ID}")
        scratch_conn.commit()
    finally:
        scratch_conn.close()

    units: List[SourceUnit] = []
    used_tables: Set[str] = set()
    for src in env_sources:
        fam = _detect_family(src)
        if fam is None:
            log.debug("Skipping unrecognized source %s", src)
            continue
        if fam == "sqlite_table":
            new_units = _ingest_sqlite(src, scratch_db, used_tables)
            units.extend(new_units)
        elif fam == "csv_table":
            new_units = _ingest_csv(src, scratch_db, used_tables)
            units.extend(new_units)
        elif fam == "excel_sheet":
            new_units = _ingest_excel(src, scratch_db, used_tables)
            units.extend(new_units)
        elif fam == "md_struct":
            new_units = _ingest_md(src, scratch_db, used_tables)
            units.extend(new_units)
    return units


__all__ = ["SourceUnit", "ingest_environment"]
